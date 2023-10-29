"""Modules required for code"""
import os

import fmpsdk
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from pydantic import ValidationError
from pydantic_functions import ApiHandling


class Globals:
    """Definition of dataframes"""

    INCOME_STATEMENT = pd.DataFrame()
    FINANCIAL_RATIOS = pd.DataFrame()


# class CheckPoint:
#     """Definition of dataframes"""

#     def check_call(self):
#         x = 10
#         return x


class AsyncAPIFunctions:
    """Class for all general API related functions"""

    async def async_load_api_keys(self) -> str | None:
        """Function to load API Key"""
        load_dotenv()
        # api_key = ApiHandling(api_key=os.getenv("FMP_API_KEY"))
        api_key = os.getenv("FMP_API_KEY")
        return api_key


class TradesOverview:
    """Obtains stock symbols for analysis

    Parameters:
            api_key (str): api key
    """

    def __init__(self, api_key: str):
        """Docstring."""
        self.api_key = api_key

    async def get_trade_symbols(self) -> list:
        """Abstract trabe symbols"""
        self.tradelist = fmpsdk.available_traded_list(apikey=self.api_key)
        # to check whether the trades list has results
        try:
            ApiHandling.check_trades_list(tradeslist=self.tradelist)
        except ValidationError as e:
            print(e)
        symbols = []
        count = 0
        for count, trade in enumerate(self.tradelist):
            if count > 10:
                break
            symbols.append(trade["symbol"])
            count += 1
        return symbols


class DataFrameFunctions:
    """Write data to dataframe

    Parameters:
            symbol (str): stock symbol
    """

    def __init__(self, symbol: str):
        """Docstring."""
        self.symbol = symbol

    def write_to_dataframe(self, data: list, name_df: str):
        """Function to write to dataframe"""
        # print(f"write to dataframe: {self.symbol}")
        df = pd.DataFrame(data)

        # Set the title of the DataFrame
        df.name = name_df

        # # Access and print the names of class variables
        # class_variable_names = list(class_dict.keys())
        # print(class_variable_names)

        # class_attributes = dir(Globals)

        # Filter and print only the class variable names
        # class_variable_names = [attr for attr in class_attributes if not callable(getattr(Globals, attr))]

        # if name_df in class_variable_names:
        #     # print(f"variable name found: {name_df}")
        #     a = class_variable_names.index(name_df)
        # print(f"Index of {name_df} in the list:", a)
        if name_df == "INCOME_STATEMENT":
            Globals.INCOME_STATEMENT = pd.concat([Globals.INCOME_STATEMENT, df], ignore_index=True)
        if name_df == "FINANCIAL_RATIOS":
            Globals.FINANCIAL_RATIOS = pd.concat([Globals.FINANCIAL_RATIOS, df], ignore_index=True)


class ExcelOperations:
    """Class to perform excel functions"""

    async def delete(sheet):
        """Deletes a given sheet"""
        # continuously delete row 2 until there
        # is only a single row left over
        # that contains column names
        while sheet.max_row > 1:
            # this method removes the row 2
            sheet.delete_rows(2)
        # return to main function
        # return

    async def clear_sheets(self):
        """Obtains all sheets for excel file"""
        # enter your file path
        path = "analytics_dataframes.xlsx"

        # load excel file
        book = openpyxl.load_workbook(path)

        # select the sheet
        sheet = book["INCOME_STATEMENT"]
        print("Maximum rows before removing:", sheet.max_row)

        await ExcelOperations.delete(sheet)
        print("Maximum rows after removing:", sheet.max_row)

        # select the sheet
        sheet = book["FINANCIAL_RATIOS"]
        print("Maximum rows before removing:", sheet.max_row)

        await ExcelOperations.delete(sheet)
        print("Maximum rows after removing:", sheet.max_row)

        # save the file to the path
        book.save(path)
        # return
